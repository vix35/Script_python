import mysql.connector
import os
# Libreria para leer archivos xlsx
import openpyxl
from openpyxl import load_workbook

# Conexion a base de datos
cnx = mysql.connector.connect(user='root',
                              password='root',
                              host='localhost',
                              database='codelco_prueba')
# Query que sirve para añadir entrada de datos de AcuAju.
add_dbf = ("INSERT INTO app_uno_dbf (numero, material, unidad, valor, fecha) VALUES (%(numero)s, %(material)s, %(unidad)s, %(valor)s, %(fecha)s)")
# Query que sirve para actualizar el valor de una entrada de datos de AcuAju.
update_dbf = ("UPDATE app_uno_dbf SET valor = %(valor)s WHERE numero = %(numero)s AND material = %(material)s AND unidad = %(unidad)s AND fecha = %(fecha)s")
# Query que sirve para revisar si una entrada de datos de AcuAju es nueva.
select_is_new_query = ("SELECT * FROM app_uno_dbf WHERE numero = %(numero)s AND  material = %(material)s AND unidad = %(unidad)s AND fecha = %(fecha)s")
# Query que sirve para verificar si una entrada de datos de AcuAju ya esta en la base de datos pero con un diferente valor.
select_is_modified_query = ("SELECT * FROM app_uno_dbf WHERE numero = %(numero)s AND  material = %(material)s AND unidad = %(unidad)s AND valor <> %(valor)s AND fecha = %(fecha)s")

# Funcion que verifica si es que ya existe una entrada de datos
def is_new(cnx, numero, material, unidad, fecha):
    mycursor = cnx.cursor(buffered=True)
    mycursor.execute(select_is_new_query,
                     {'numero': numero, 'material': material, 'unidad': unidad, 'fecha': fecha})
    row_count = mycursor.rowcount
    mycursor.close()
    if row_count >= 1:
        return False
    else:
        return True

# Funcion que verifica si es que la entrada de datos existente es diferente a la que se quiere ingresar
def is_modified(cnx, numero, material, unidad, valor, fecha):
    mycursor = cnx.cursor(buffered=True)
    mycursor.execute(select_is_modified_query,
                     {'numero': numero, 'material': material, 'unidad': unidad, 'valor': valor, 'fecha': fecha})
    row_count = mycursor.rowcount
    mycursor.close()
    if row_count >= 1:
        return True
    else:
        return False

# Funcion que ingresa entrada de datos a base de datos.
def add(cnx, numero, material, unidad, valor, fecha):
    mycursor = cnx.cursor(buffered=True)
    mycursor.execute(add_dbf,{'numero': numero, 'material': material, 'unidad': unidad, 'valor': valor, 'fecha': fecha})
    cnx.commit()
    mycursor.close()

# FUncion que actualiza entrada de datos existente en la base de datos.
def update(cnx, numero, material, unidad, valor, fecha):
    mycursor = cnx.cursor(buffered=True)
    mycursor.execute(update_dbf, {'numero': numero, 'material': material, 'unidad': unidad, 'valor': valor, 'fecha': fecha})
    cnx.commit()
    mycursor.close()

# Funcion que lee los archivos y analiza lineas de datos.
def save_data(archivo, path):
    dir = os.path.join(path, archivo)
    f = open(dir)
    wb = load_workbook(filename=dir)
    sheet = wb.get_sheet_by_name('Hoja1')

    for i, linea in enumerate(sheet.iter_rows()):
        if i == 0:
            pass
        else:
            if is_new(  cnx,
                        linea[0].internal_value,
                        linea[1].internal_value,
                        linea[2].internal_value,
                        linea[4].internal_value):
                add(cnx,
                    linea[0].internal_value,
                    linea[1].internal_value,
                    linea[2].internal_value,
                    linea[3].internal_value,
                    linea[4].internal_value)
            else:
                if is_modified( cnx,
                                linea[0].internal_value,
                                linea[1].internal_value,
                                linea[2].internal_value,
                                linea[3].internal_value,
                                linea[4].internal_value):
                    update( cnx,
                            linea[0].internal_value,
                            linea[1].internal_value,
                            linea[2].internal_value,
                            linea[3].internal_value,
                            linea[4].internal_value)
                else:
                    continue
    f.close()

#Direccion en donde se encuentran los archivos a actualizar en la base de datos.
path = r"C:\Users\vicen\Documents\dbf"
entries = os.listdir(path)

for entry in entries:
    save_data(entry, path)
